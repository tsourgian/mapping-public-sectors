from openpyxl import load_workbook
import codecs
file = open("result.txt", "w",encoding="utf-8")
wb = load_workbook('test.xlsx')
ws2=wb['ΦΟΡΕΙΣ']
#[0]"ΟΝΟΜΑΣΙΑ ΦΟΡΕΑ - ΥΠΗΡΕΣΙΑΣ",[1]"ΕΠΟΠΤΕΥΟΝ ΥΠΟΥΡΓΕΙΟ",[2]"ΝΟΜΙΚΗ ΜΟΡΦΗ - ΕΙΔΟΣ ΥΠΗΡΕΣΙΑΣ",[3]"ΤΟΜΕΙΣ ΠΟΛΙΤΙΚΗΣ",[4]"ΥΠΑΓΩΓΗ ΣΤΟ ΔΗΜΟΣΙΟ ΤΟΜΕΑ",[5]"ΛΟΙΠΕΣ ΠΛΗΡΟΦΟΡΙΕΣ"
for row_cells in ws2.iter_rows():
    if row_cells[0].value==None :
     row_cells[0].value=""
    if row_cells[1].value == None:
         row_cells[1].value = ""
    if row_cells[2].value == None:
             row_cells[2].value = ""
    if row_cells[3].value == None:
             row_cells[3].value = ""
    if row_cells[4].value==None :
     row_cells[4].value=""
    if row_cells[5].value == None:
         row_cells[5].value = ""
    file.write("\""+row_cells[0].value+"\""+",\""+row_cells[1].value+"\""+",\""+row_cells[2].value+"\""",\""+row_cells[3].value+"\""",\""+row_cells[4].value+"\""",\""+row_cells[5].value+"\"\n")

file.close()
